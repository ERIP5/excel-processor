from pathlib import Path

import openpyxl
import pandas as pd

### Global variables
data = Path("data")  # Chemin vers le dossier data

colonnes = [
    "Identifiant",
    "Abréviation",
    "Code PPN",
    "Nom complet",
    "Promotion",
    "Période",
    "Responsable",
]

### Aux. functions


def getPathwaysRange(moduleSheet):
    currentPathway = moduleSheet[f"C{1}"].value
    pathways = {currentPathway: [1]}

    for i in range(1, moduleSheet.max_row + 1):
        if moduleSheet[f"C{i}"].value != currentPathway:
            pathways[currentPathway].append(i - 1)
            currentPathway = moduleSheet[f"C{i}"].value
            pathways[currentPathway] = [i]

    pathways[currentPathway].append(moduleSheet.max_row)

    return pathways


def extractModuleInformation(moduleExcelLine: tuple):
    # Return a single row DataFrame containing information regarding a specific module

    moduleLine = moduleExcelLine[0]
    moduleDataFrame = pd.DataFrame(columns=colonnes)

    # Identfiant, Abréviation, Code PPN, Nom complet

    rawData = {"Identifiant": moduleLine[0].value, "Abréviation": moduleLine[1].value}

    rawDataFrame = pd.DataFrame(
        rawData, index=[0]
    )  # On créé un dataframe contenant les données bruts

    moduleDataFrame = pd.concat([moduleDataFrame, rawDataFrame]).reset_index(
        drop=True
    )  # On l'ajoute au dataframe principal

    moduleDataFrame["Code PPN"] = moduleDataFrame["Identifiant"]
    moduleDataFrame["Nom complet"] = moduleDataFrame["Abréviation"]

    # Période et Promotion

    description = moduleLine[11].value

    # print(description)

    descriptionSegmented = description.replace("/", ",").split(",")

    # print(descriptionSegmented)

    # Removing trailing spaces and empty strings
    indexToRemove = []

    for i in range(len(descriptionSegmented)):
        processedString = descriptionSegmented[i].strip()

        if processedString == "":
            indexToRemove.append(i)
        else:
            descriptionSegmented[i] = processedString

    for index in reversed(indexToRemove):
        descriptionSegmented.pop(index)

    semesterSegment = descriptionSegmented[1]
    # print(semesterSegment.strip("Semestre").strip())
    semesterNumber = semesterSegment.strip("Semestre").strip()[1]

    moduleDataFrame["Période"] = f"S{semesterNumber}"

    year = int(semesterNumber) // 2 + 1  # 2 semestres par année

    pathway = moduleLine[2].value

    yearGroup = f"{pathway}{year}"
    # print(yearGroup)

    moduleDataFrame["Promotion"] = yearGroup

    # Responsable

    # TODO: avec le nom et prénom du responsable, il faut pouvoir obtenir son identifiant qui peut ne pas être ses initiales

    tutorIdentity = moduleLine[4].value
    tutorInitials = "".join([string[0].upper() for string in tutorIdentity.split(" ")])
    moduleDataFrame["Responsable"] = tutorInitials

    return moduleDataFrame


def batchModuleInformation(excelSheet, rowRange):
    modulesDF = pd.DataFrame(columns=colonnes)

    lastColumnLetter = f"{openpyxl.utils.cell.get_column_letter(excelSheet.max_column)}"

    for rowNumber in range(rowRange[0], rowRange[1]):
        lastColumnCoordinates = f"{lastColumnLetter}{rowNumber}"
        processedLine = excelSheet[f"A{rowNumber}":lastColumnCoordinates]
        processedDataFrame = extractModuleInformation(processedLine)

        modulesDF = pd.concat([modulesDF, processedDataFrame]).reset_index(drop=True)

    return modulesDF


def DataFramesToExcel(
    dataframes: [pd.DataFrame], sheetNames: [str], output: str or Path
):
    # Create an Excel file in which each sheet is associated with a dataframe in the list "dataframes" and its name is indicated in the list "sheetNames"

    with pd.ExcelWriter(output) as writer:
        for index in range(len(sheetNames)):
            dataframes[index].to_excel(
                writer, sheet_name=sheetNames[index], index=False
            )


### Main


def main(moduleExcelPath):
    moduleWorkbook = openpyxl.load_workbook(moduleExcelPath)
    moduleSheet = moduleWorkbook[moduleWorkbook.sheetnames[0]]

    pathwaysRange = getPathwaysRange(moduleSheet)
    selectedPathway = "INFO"
    pathwayRange = pathwaysRange[selectedPathway]

    moduleDataFrame = batchModuleInformation(moduleSheet, pathwayRange)
    DataFramesToExcel([moduleDataFrame], ["Modules"], Path("output/modules.xlsx"))


main(data / "liste_intervenants.xlsx")