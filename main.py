import openpyxl

with open("modulesExtractor.py", "r") as fichier:
    contenu = fichier.read()
    exec(contenu)

with open("intervenantsExtractor.py", "r") as fichier1:
    contenu1 = fichier1.read()
    exec(contenu1)

# Load the first workbook
wb1 = openpyxl.load_workbook('output/modules.xlsx')

# Load the second workbook
wb2 = openpyxl.load_workbook('output/intervenants.xlsx')

# Create a new workbook
wb3 = openpyxl.load_workbook('data/database_file.xlsx')

# Create a new workbook
wb4 = openpyxl.Workbook()

# Copy sheets from the first workbook to the new workbook
for sheet in wb1:
    new_sheet = wb4.create_sheet(title="{}".format(sheet.title))
    for row in sheet:
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        new_sheet.append(new_row)

# Copy sheets from the second workbook to the new workbook
for sheet in wb2:
    new_sheet = wb4.create_sheet(title="{}".format(sheet.title))
    for row in sheet:
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        new_sheet.append(new_row)

# Copy sheets from the second workbook to the new workbook
for sheet in wb3:
    new_sheet = wb4.create_sheet(title="{}".format(sheet.title))
    for row in sheet:
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        new_sheet.append(new_row)



# Remove the default first sheet from the new workbook
wb4.remove(wb4['Sheet'])

# Save the new workbook
wb4.save('final.xlsx')